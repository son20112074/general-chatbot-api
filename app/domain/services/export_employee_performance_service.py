from datetime import datetime
import os
import pandas as pd
from typing import List, Dict, Any
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet

from app.core.config import settings


class EmployeePerformanceExcelCreator:
    def __init__(self, session: AsyncSession):
        self.session = session
    
    async def create_excel(self, employees_data: List[Dict[str, Any]], period_info: Dict[str, Any]) -> str:
        """Create Excel file with employee performance data"""
        try:
            # Create DataFrame from employee data
            df = pd.DataFrame(employees_data)
            
            # Rename columns to Vietnamese
            column_mapping = {
                'user_id': 'Mã người dùng',
                'account_name': 'Tên tài khoản',
                'full_name': 'Họ và tên',
                'email': 'Email',
                'total_tasks': 'Tổng số nhiệm vụ',
                'completed_tasks': 'Nhiệm vụ hoàn thành',
                'completion_rate': 'Tỷ lệ hoàn thành (%)'
            }
            
            df = df.rename(columns=column_mapping)
            
            # Add ranking column
            df.insert(0, 'Xếp hạng', range(1, len(df) + 1))
            
            # Format completion rate to 1 decimal place
            if 'Tỷ lệ hoàn thành (%)' in df.columns:
                df['Tỷ lệ hoàn thành (%)'] = df['Tỷ lệ hoàn thành (%)'].round(1)
            
            # Create file name with timestamp
            file_name = f'bao_cao_hieu_suat_nhan_vien_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
            
            # Create save path with year/month/day structure
            current_date = datetime.now()
            year = str(current_date.year)
            month = str(current_date.month).zfill(2)
            day = str(current_date.day).zfill(2)
            
            save_path = f'static/downloads/excel/{year}/{month}/{day}'
            
            # Create directory structure if it doesn't exist
            os.makedirs(save_path, exist_ok=True)
            
            file_path = os.path.join(save_path, file_name)
            
            # Create workbook manually for better formatting control
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create main data sheet with formatting
            self._create_main_sheet(wb, df, period_info)
            
            # Create summary sheet
            self._create_summary_sheet(wb, df, period_info)
            
            # Create statistics sheet
            self._create_statistics_sheet(wb, df, period_info)
            
            # Save workbook
            wb.save(file_path)
            
            return settings.BASE_URL + '/' + save_path + '/' + file_name
            
        except Exception as e:
            raise Exception(f"Error creating Excel file: {str(e)}")
    
    def _create_main_sheet(self, wb: Workbook, df: pd.DataFrame, period_info: Dict[str, Any]):
        """Create main data sheet with beautiful formatting"""
        ws = wb.create_sheet("Báo cáo hiệu suất")
        
        # Define styles
        title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        subtitle_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        normal_font = Font(name='Arial', size=10)
        
        title_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        subtitle_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        alt_row_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add title
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = "BÁO CÁO HIỆU SUẤT NHÂN VIÊN"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = center_alignment
        title_cell.border = thin_border
        
        # Add subtitle with period info
        ws.merge_cells('A2:G2')
        subtitle_cell = ws['A2']
        from_time = period_info.get('from_time', 'N/A')
        to_time = period_info.get('to_time', 'N/A')
        subtitle_cell.value = f"Thời gian: {from_time} - {to_time}"
        subtitle_cell.font = subtitle_font
        subtitle_cell.fill = subtitle_fill
        subtitle_cell.alignment = center_alignment
        subtitle_cell.border = thin_border
        
        # Add summary info
        ws.merge_cells('A3:G3')
        summary_cell = ws['A3']
        total_employees = len(df)
        total_tasks = df['Tổng số nhiệm vụ'].sum() if 'Tổng số nhiệm vụ' in df.columns else 0
        total_completed = df['Nhiệm vụ hoàn thành'].sum() if 'Nhiệm vụ hoàn thành' in df.columns else 0
        summary_cell.value = f"Tổng số nhân viên: {total_employees} | Tổng số nhiệm vụ: {total_tasks} | Nhiệm vụ hoàn thành: {total_completed}"
        summary_cell.font = subtitle_font
        summary_cell.fill = subtitle_fill
        summary_cell.alignment = center_alignment
        summary_cell.border = thin_border
        
        # Add empty row
        ws.row_dimensions[4].height = 10
        
        # Add headers
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=5, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        
        # Add data rows
        for row_idx, (_, row_data) in enumerate(df.iterrows(), 6):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.font = normal_font
                cell.alignment = center_alignment if col_idx <= 2 else left_alignment
                cell.border = thin_border
                
                # Alternate row colors
                if row_idx % 2 == 0:
                    cell.fill = alt_row_fill
        
        # Set column widths - simple approach to avoid merged cell issues
        default_widths = [8, 15, 25, 20, 15, 15, 15]  # For ranking, ID, name, email, total, completed, rate
        for i, width in enumerate(default_widths[:len(df.columns)]):
            try:
                column_letter = ws.cell(row=5, column=i+1).column_letter
                ws.column_dimensions[column_letter].width = width
            except:
                pass
        
        # Set row heights
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 25
        ws.row_dimensions[3].height = 25
        ws.row_dimensions[5].height = 25
    
    def _create_summary_sheet(self, wb: Workbook, df: pd.DataFrame, period_info: Dict[str, Any]):
        """Create summary sheet with key metrics"""
        ws = wb.create_sheet("Tóm tắt")
        
        # Define styles
        title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        normal_font = Font(name='Arial', size=10)
        bold_font = Font(name='Arial', size=10, bold=True)
        
        title_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        section_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add title
        ws.merge_cells('A1:B1')
        title_cell = ws['A1']
        title_cell.value = "TÓM TẮT BÁO CÁO HIỆU SUẤT"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = center_alignment
        title_cell.border = thin_border
        
        # Add headers
        ws['A2'] = 'Chỉ số'
        ws['B2'] = 'Giá trị'
        for col in ['A2', 'B2']:
            cell = ws[col]
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        
        current_row = 3
        
        # Basic statistics
        basic_stats = [
            ['Tổng số nhân viên', len(df)],
            ['Tổng số nhiệm vụ', df['Tổng số nhiệm vụ'].sum() if 'Tổng số nhiệm vụ' in df.columns else 0],
            ['Tổng số nhiệm vụ hoàn thành', df['Nhiệm vụ hoàn thành'].sum() if 'Nhiệm vụ hoàn thành' in df.columns else 0],
            ['Tỷ lệ hoàn thành trung bình (%)', round(df['Tỷ lệ hoàn thành (%)'].mean(), 1) if 'Tỷ lệ hoàn thành (%)' in df.columns else 0]
        ]
        
        for stat in basic_stats:
            ws[f'A{current_row}'] = stat[0]
            ws[f'B{current_row}'] = stat[1]
            for col in ['A', 'B']:
                cell = ws[f'{col}{current_row}']
                cell.font = normal_font
                cell.alignment = left_alignment
                cell.border = thin_border
            current_row += 1
        
        # Empty row
        current_row += 1
        
        # Period information section
        ws.merge_cells(f'A{current_row}:B{current_row}')
        period_header = ws[f'A{current_row}']
        period_header.value = 'THÔNG TIN THỜI GIAN'
        period_header.font = bold_font
        period_header.fill = section_fill
        period_header.alignment = center_alignment
        period_header.border = thin_border
        current_row += 1
        
        period_info_data = [
            ['Từ ngày', period_info.get('from_time', 'N/A')],
            ['Đến ngày', period_info.get('to_time', 'N/A')],
            ['Tổng số user được phân tích', period_info.get('total_users_analyzed', 'N/A')]
        ]
        
        for info in period_info_data:
            ws[f'A{current_row}'] = info[0]
            ws[f'B{current_row}'] = info[1]
            for col in ['A', 'B']:
                cell = ws[f'{col}{current_row}']
                cell.font = normal_font
                cell.alignment = left_alignment
                cell.border = thin_border
            current_row += 1
        
        # Empty row
        current_row += 1
        
        # Performance categories section
        ws.merge_cells(f'A{current_row}:B{current_row}')
        perf_header = ws[f'A{current_row}']
        perf_header.value = 'PHÂN LOẠI HIỆU SUẤT'
        perf_header.font = bold_font
        perf_header.fill = section_fill
        perf_header.alignment = center_alignment
        perf_header.border = thin_border
        current_row += 1
        
        if 'Tỷ lệ hoàn thành (%)' in df.columns:
            excellent = len(df[df['Tỷ lệ hoàn thành (%)'] >= 90])
            good = len(df[(df['Tỷ lệ hoàn thành (%)'] >= 70) & (df['Tỷ lệ hoàn thành (%)'] < 90)])
            average = len(df[(df['Tỷ lệ hoàn thành (%)'] >= 50) & (df['Tỷ lệ hoàn thành (%)'] < 70)])
            below_average = len(df[df['Tỷ lệ hoàn thành (%)'] < 50])
            
            perf_categories = [
                ['Xuất sắc (≥90%)', excellent],
                ['Tốt (70-89%)', good],
                ['Trung bình (50-69%)', average],
                ['Dưới trung bình (<50%)', below_average]
            ]
            
            for category in perf_categories:
                ws[f'A{current_row}'] = category[0]
                ws[f'B{current_row}'] = category[1]
                for col in ['A', 'B']:
                    cell = ws[f'{col}{current_row}']
                    cell.font = normal_font
                    cell.alignment = left_alignment
                    cell.border = thin_border
                current_row += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        
        # Set row heights
        ws.row_dimensions[1].height = 30
        ws.row_dimensions[2].height = 25
    
    def _create_statistics_sheet(self, wb: Workbook, df: pd.DataFrame, period_info: Dict[str, Any]):
        """Create statistics sheet with detailed analysis"""
        ws = wb.create_sheet("Thống kê chi tiết")
        
        # Define styles
        title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        normal_font = Font(name='Arial', size=10)
        bold_font = Font(name='Arial', size=10, bold=True)
        
        title_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        section_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add title
        ws.merge_cells('A1:B1')
        title_cell = ws['A1']
        title_cell.value = "THỐNG KÊ CHI TIẾT HIỆU SUẤT"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = center_alignment
        title_cell.border = thin_border
        
        current_row = 2
        
        # Top performers section
        ws.merge_cells(f'A{current_row}:B{current_row}')
        top_header = ws[f'A{current_row}']
        top_header.value = 'TOP 5 NHÂN VIÊN XUẤT SẮC NHẤT'
        top_header.font = bold_font
        top_header.fill = section_fill
        top_header.alignment = center_alignment
        top_header.border = thin_border
        current_row += 1
        
        if 'Tỷ lệ hoàn thành (%)' in df.columns:
            top_5 = df.nlargest(5, 'Tỷ lệ hoàn thành (%)')
            for idx, row in top_5.iterrows():
                ws[f'A{current_row}'] = f"{row['Xếp hạng']}. {row['Họ và tên']}"
                ws[f'B{current_row}'] = f"{row['Tỷ lệ hoàn thành (%)']}% ({row['Nhiệm vụ hoàn thành']}/{row['Tổng số nhiệm vụ']})"
                for col in ['A', 'B']:
                    cell = ws[f'{col}{current_row}']
                    cell.font = normal_font
                    cell.alignment = left_alignment
                    cell.border = thin_border
                current_row += 1
        
        # Empty row
        current_row += 1
        
        # Performance distribution section
        ws.merge_cells(f'A{current_row}:B{current_row}')
        dist_header = ws[f'A{current_row}']
        dist_header.value = 'PHÂN BỐ HIỆU SUẤT'
        dist_header.font = bold_font
        dist_header.fill = section_fill
        dist_header.alignment = center_alignment
        dist_header.border = thin_border
        current_row += 1
        
        if 'Tỷ lệ hoàn thành (%)' in df.columns:
            # Create performance ranges
            ranges = [
                ('90-100%', 90, 100),
                ('80-89%', 80, 89),
                ('70-79%', 70, 79),
                ('60-69%', 60, 69),
                ('50-59%', 50, 59),
                ('40-49%', 40, 49),
                ('30-39%', 30, 39),
                ('20-29%', 20, 29),
                ('10-19%', 10, 19),
                ('0-9%', 0, 9)
            ]
            
            for range_name, min_val, max_val in ranges:
                count = len(df[(df['Tỷ lệ hoàn thành (%)'] >= min_val) & (df['Tỷ lệ hoàn thành (%)'] <= max_val)])
                if count > 0:
                    ws[f'A{current_row}'] = range_name
                    ws[f'B{current_row}'] = count
                    for col in ['A', 'B']:
                        cell = ws[f'{col}{current_row}']
                        cell.font = normal_font
                        cell.alignment = left_alignment
                        cell.border = thin_border
                    current_row += 1
        
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 30
        
        # Set row heights
        ws.row_dimensions[1].height = 30
    
    async def get_all_employee_performance(self, user_ids: List[int], from_time: datetime, to_time: datetime) -> List[Dict[str, Any]]:
        """Get performance data for all employees without limit"""
        try:
            # Query to get task statistics per user (without LIMIT)
            performance_query = text("""
                SELECT 
                    u.id as user_id,
                    u.full_name,
                    u.email,
                    u.account_name,
                    COUNT(t.id) as total_tasks,
                    COUNT(CASE WHEN t.status = 'completed' THEN 1 END) as completed_tasks,
                    CASE 
                        WHEN COUNT(t.id) > 0 THEN 
                            ROUND((COUNT(CASE WHEN t.status = 'completed' THEN 1 END) * 100.0 / COUNT(t.id)), 1)
                        ELSE 0 
                    END as completion_rate
                FROM users u
                LEFT JOIN tasks t ON u.id = ANY(t.assigned_to)
                    AND t.created_at >= :from_time 
                    AND t.created_at <= :to_time
                WHERE u.id = ANY(:user_ids)
                    AND u.status = true  -- Only include active users (not deleted)
                GROUP BY u.id, u.full_name, u.email, u.account_name
                ORDER BY completion_rate DESC, completed_tasks DESC
            """)
            
            result = await self.session.execute(
                performance_query,
                {
                    "user_ids": user_ids,
                    "from_time": from_time,
                    "to_time": to_time
                }
            )

            employees = []
            for row in result:
                employees.append({
                    'user_id': row[0],
                    'full_name': row[1] or 'Unknown',
                    'email': row[2],
                    'account_name': row[3],
                    'total_tasks': row[4],
                    'completed_tasks': row[5],
                    'completion_rate': float(row[6]) if row[6] else 0.0
                })
            
            return employees
            
        except Exception as e:
            raise Exception(f"Error getting employee performance: {str(e)}") 